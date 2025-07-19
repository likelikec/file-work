import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import classification_report, roc_auc_score, precision_recall_curve, f1_score
from sklearn.preprocessing import LabelEncoder
import matplotlib

matplotlib.use('Agg')  # 必须在其他matplotlib导入前设置
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import RocCurveDisplay
import shap  # 新增SHAP库导入
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

# 过滤Matplotlib非GUI后端警告
warnings.filterwarnings("ignore", category=UserWarning, module="matplotlib")

# 设置中文字体支持（根据系统实际可用字体调整）
plt.rcParams["font.family"] = ['Microsoft YaHei', 'Microsoft JhengHei', 'SimHei', ]
plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

# 读取数据
df = pd.read_excel(r"C:\Users\17958\Desktop\symtrain-01.xlsx", engine='openpyxl')

# 预处理数据
# 删除因变量为2的行（第10列，索引9）
df = df[df['1适合LLM'] != 2]
df['1适合LLM'] = df['1适合LLM'].astype(int)

# 定义特征和标签
features = ['B', 'COM_RAT', 'Cyclic', 'D',
            'Dcy*', 'DIT', 'DPT*', 'E', 'Inner', 'LCOM', 'Level', 'LOC', 'N',
            'NCLOC', 'NOAC', 'NOC', 'NOIC', 'OCmax', 'PDcy', 'PDpt', 'STAT',
            'TCOM_RAT', 'V', 'WMC', 'CBO', 'CLOC', 'Command', 'CONS', 'CSA', 'CSO',
            'CSOA', 'Dcy', 'DPT', 'INNER', 'jf', 'JLOC', 'Jm', 'Level*', 'MPC', 'n',
            'NAAC', 'NAIC', 'NOOC', 'NTP', 'OCavg', 'OPavg', 'OSavg', 'OSmax',
            'Query', 'RFC', 'TODO', "String processing", "File operations", "Network communication",
            "Database operations", "Mathematical calculation", "User Interface",
            "Business Logic", "Data Structures and Algorithms", "Systems and Tools",
            "Concurrency and Multithreading", "Exception handling"]
X = df[features]
y = df['1适合LLM']
# 使用0填充缺失值
X = X.fillna(0)
y = y.fillna(0)

# 划分训练集和测试集（分层抽样）
X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.2, stratify=y, random_state=42)

# 参数网格
param_grid = {
    'n_estimators': range(50, 120, 5),
    'max_depth': [None, 3, 5, 8, 10],
    'min_samples_split': [2, 3, 5, 8, 10],
    'class_weight': ['balanced', None]
}

# 初始化随机森林
rf = RandomForestClassifier(random_state=42)

# 网格搜索（使用roc_auc作为评估指标）
grid_search = GridSearchCV(
    estimator=rf,
    param_grid=param_grid,
    scoring='roc_auc',
    cv=5,
    n_jobs=-1
)

grid_search.fit(X_train, y_train)

# 获取最佳模型
best_model = grid_search.best_estimator_

# 阈值优化（使用验证集）
X_train_sub, X_val, y_train_sub, y_val = train_test_split(
    X_train, y_train, test_size=0.2, stratify=y_train, random_state=42)
best_model.fit(X_train_sub, y_train_sub)
y_proba_val = best_model.predict_proba(X_val)[:, 1]

# 寻找最佳阈值
precisions, recalls, thresholds = precision_recall_curve(y_val, y_proba_val)
f1_scores = 2 * (precisions * recalls) / (precisions + recalls + 1e-8)
best_threshold_idx = np.argmax(f1_scores[:-1])  # 排除最后一个元素
best_threshold = thresholds[best_threshold_idx]

# 应用最佳阈值到测试集
y_proba_test = best_model.predict_proba(X_test)[:, 1]
y_pred_test = (y_proba_test >= best_threshold).astype(int)

# 输出结果
print("最佳参数组合：")
print(grid_search.best_params_)
print("\n最佳分类阈值：", round(best_threshold, 3))
print("\n测试集分类报告：")
print(classification_report(y_test, y_pred_test))
print("AUC-ROC:", roc_auc_score(y_test, y_proba_test))

# 输出特征重要性
importances = best_model.feature_importances_
print("\n特征重要性：")
for idx, imp in enumerate(importances):
    print(f"特征 {X.columns[idx]}: {imp:.3f}")

# 绘制ROC曲线
plt.figure(figsize=(8, 6))
RocCurveDisplay.from_estimator(
    best_model,
    X_test,
    y_test,
    name='随机森林',
    color='darkorange'
)
plt.plot([0, 1], [0, 1], 'k--', lw=2)  # 添加对角线
plt.title('ROC曲线 (测试集)')
plt.grid(alpha=0.3)
plt.box(False)
plt.savefig('roc_curve.png', dpi=300)  # 保存图片
plt.close()  # 关闭图形

# 绘制特征重要性条形图
plt.figure(figsize=(12, 8))

# 创建特征重要性DataFrame并排序（包含所有特征）
feature_importance_df = pd.DataFrame({
    '特征': X.columns,
    '重要性': importances
})
feature_importance_df = feature_importance_df.sort_values('重要性', ascending=False)


# ====== 新增：将特征重要性保存为Excel表格 ======
def save_feature_importance_to_excel(df, file_path):
    """将特征重要性DataFrame保存为格式化的Excel文件"""
    wb = Workbook()
    ws = wb.active

    # 设置表头样式
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='4285F4', end_color='4285F4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # 写入表头
    ws.append(df.columns.tolist())
    header_row = ws[1]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入数据行并设置样式
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    # 设置数据列样式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            # 重要性数值保留3位小数
            if cell.column == 2:
                cell.number_format = '0.000'
            # 交替行背景色提高可读性
            if (cell.row - 1) % 2 == 0:
                cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            cell.alignment = Alignment(horizontal='left' if cell.column == 1 else 'center', vertical='center')

    # 自动调整列宽
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = length + 2

    # 保存文件
    wb.save(file_path)
    print(f"特征重要性已保存至: {file_path}")


# 保存特征重要性到Excel
feature_importance_file = 'feature_importance.xlsx'
save_feature_importance_to_excel(feature_importance_df, feature_importance_file)

# 绘制特征重要性前30的条形图
plt.figure(figsize=(12, 8))
top_30_features = feature_importance_df.head(30)
sns.barplot(x='重要性', y='特征', data=top_30_features, palette='viridis')
plt.title('特征重要性排名（前30）')
plt.xlabel('重要性')
plt.ylabel('特征')
plt.tight_layout()
plt.savefig('feature_importance.png', dpi=300, bbox_inches='tight')
plt.close()  # 关闭图形
print("\n已生成特征重要性条形图：feature_importance.png")

### ====== 新增：SHAP值计算与可视化 ======
print("\n开始计算SHAP值...")

# 1. 计算SHAP值（使用训练子集加速计算）
# 为避免计算量过大，使用训练集的前100个样本作为背景
background = X_train_sub.sample(n=100, random_state=42)
explainer = shap.Explainer(best_model)
shap_values = explainer(background)

# 2. 全局解释：SHAP摘要图（展示所有特征的影响分布）
plt.figure(figsize=(12, 8))
shap.summary_plot(shap_values.values, background, feature_names=X.columns,
                  plot_type="bar", show=False)
plt.title('SHAP特征重要性摘要（条形图）')
plt.tight_layout()
plt.savefig('shap_summary_bar.png', dpi=300)
plt.close()  # 关闭图形

plt.figure(figsize=(12, 8))
shap.summary_plot(shap_values.values, background, feature_names=X.columns,
                  plot_type="dot", show=False)
plt.title('SHAP特征重要性摘要（点图）')
plt.tight_layout()
plt.savefig('shap_summary_dot.png', dpi=300)
plt.close()  # 关闭图形

# 3. 局部解释：单个样本的SHAP力导向图
# 选择测试集中的一个样本进行解释
sample_idx = 0  # 可修改为任意样本索引
plt.figure(figsize=(12, 6))
# 修复：使用新版SHAP waterfall图API
shap.plots.waterfall(shap_values[sample_idx][:, 1], max_display=40)
plt.title(f'样本 {sample_idx} 的SHAP力导向图')
plt.tight_layout()
plt.savefig(f'shap_waterfall_sample{sample_idx}.png', dpi=300)
plt.close()  # 关闭图形

# 4. 特征依赖关系：SHAP依赖图
# 选择最重要的特征绘制依赖关系
top_feature = feature_importance_df.iloc[0]['特征']
plt.figure(figsize=(10, 6))
# 修复：正确传递单个特征列，使用值的第一个维度
shap_feature_values = shap_values[:, top_feature].values[:, 0]
shap.plots.scatter(shap_values[:, top_feature], color=shap_feature_values)
plt.title(f'{top_feature} 的SHAP依赖关系图')
plt.tight_layout()
plt.savefig(f'shap_dependence_{top_feature}.png', dpi=300)
plt.close()  # 关闭图形

# 5. 特征重要性对比：SHAP与模型原生重要性
plt.figure(figsize=(12, 8))
shap.plots.bar(shap_values, max_display=40, show=False)
plt.title('SHAP特征重要性（前62）')
plt.tight_layout()
plt.savefig('shap_feature_importance.png', dpi=300)
plt.close()  # 关闭图形

print("SHAP值计算与可视化完成，结果已保存至：")
print("1. shap_summary_bar.png - SHAP特征重要性摘要（条形图）")
print("2. shap_summary_dot.png - SHAP特征重要性摘要（点图）")
print(f"3. shap_waterfall_sample{sample_idx}.png - 单样本力导向图")
print(f"4. shap_dependence_{top_feature}.png - 特征依赖关系图")
print("5. shap_feature_importance.png - SHAP特征重要性排序")
